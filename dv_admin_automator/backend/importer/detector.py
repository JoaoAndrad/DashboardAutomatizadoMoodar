from typing import Optional ,Dict ,Any ,List 
import os 
import csv 
import re 
def _try_read_with_pandas (path :str ):
    try :
        import pandas as pd 
    except Exception :
        return None 
    ext =os .path .splitext (path )[1 ].lower ()
    try :
        if ext in ('.xlsx','.xls'):
            df =pd .read_excel (path )
        elif ext =='.csv':
            for enc in ('utf-8','latin-1','cp1252'):
                try :
                    df =pd .read_csv (path ,encoding =enc )
                    break 
                except Exception :
                    df =None 
            if df is None :
                return None 
        else :
            df =pd .read_csv (path ,sep ='\t')
        return df 
    except Exception :
        return None 
def _read_xlsx_with_openpyxl (path :str ,max_rows :int =20 ):
    try :
        from openpyxl import load_workbook 
    except Exception :
        raise ImportError ("openpyxl not installed; install pandas or openpyxl to read XLSX files")
    try :
        wb =load_workbook (filename =path ,read_only =True ,data_only =True )
        ws =wb .active 
        rows =[]
        count =0 
        for row in ws .iter_rows (values_only =True ):
            rows .append ([(''if v is None else str (v ))for v in row ])
            count +=1 
            if count >=max_rows +1 :
                break 
        if not rows :
            return {'columns':[],'rows':[]}
        header =[str (h )for h in rows [0 ]]
        preview_rows =[]
        for r in rows [1 :max_rows +1 ]:
            d ={}
            for i ,col in enumerate (header ):
                d [col ]=r [i ]if i <len (r )else None 
            preview_rows .append (d )
        return {'columns':header ,'rows':preview_rows }
    except Exception :
        return {'columns':[],'rows':[]}
def _fallback_read_preview (path :str ,max_rows :int =20 ):
    rows :List [List [str ]]=[]
    columns :List [str ]=[]
    try :
        with open (path ,'r',encoding ='utf-8',errors ='ignore')as f :
            sample =f .read (8192 )
            sep ='\t'if '\t'in sample and sample .count ('\t')>sample .count (',')else ','
        with open (path ,'r',encoding ='utf-8',errors ='ignore')as f :
            reader =csv .reader (f ,delimiter =sep )
            for r in reader :
                rows .append (r )
                if len (rows )>=max_rows +1 :
                    break 
        if not rows :
            return {'columns':[],'rows':[]}
        header =rows [0 ]
        columns =[str (h )for h in header ]
        preview_rows =[]
        for r in rows [1 :max_rows +1 ]:
            d ={}
            for i ,col in enumerate (columns ):
                d [col ]=r [i ]if i <len (r )else None 
            preview_rows .append (d )
        return {'columns':columns ,'rows':preview_rows }
    except Exception :
        return {'columns':[],'rows':[]}
def detect_file_type (path :str ,sample_rows :int =20 )->Dict [str ,Any ]:
    result :Dict [str ,Any ]={'type':None ,'candidates':{'cpf':None ,'email':None ,'name':None },'preview':{'columns':[],'rows':[]}}
    df =_try_read_with_pandas (path )
    if df is not None :
        try :
            cols =list (df .columns .astype (str ))
            preview_rows =[]
            for _ ,r in df .head (sample_rows ).iterrows ():
                preview_rows .append ({c :(None if r [c ]is None else str (r [c ]))for c in cols })
            result ['preview']={'columns':cols ,'rows':preview_rows }
            cpf_re =re .compile (r'^\s*\d{3}\.??\d{3}\.??\d{3}-?\d{2}\s*$')
            for col in cols :
                sample =[str (x )for x in df [col ].dropna ().astype (str ).head (20 )]
                if not result ['candidates']['cpf']:
                    if any (cpf_re .match (s )for s in sample ):
                        result ['candidates']['cpf']=col 
                if not result ['candidates']['email']:
                    if any ('@'in s for s in sample ):
                        result ['candidates']['email']=col 
                if not result ['candidates']['name']:
                    if any (len (s .split ())>=2 for s in sample ):
                        result ['candidates']['name']=col 
            if result ['candidates']['cpf']and result ['candidates']['name']:
                result ['type']='cpf'
            elif result ['candidates']['email']:
                result ['type']='email'
            else :
                result ['type']=None 
            return result 
        except Exception :
            pass 
    ext =os .path .splitext (path )[1 ].lower ()
    if df is None and ext in ('.xlsx','.xls'):
        try :
            preview =_read_xlsx_with_openpyxl (path ,max_rows =sample_rows )
            result ['preview']=preview 
            cols =preview .get ('columns',[])
            for col in cols :
                sample =[str (r .get (col ,''))for r in preview .get ('rows',[])if r .get (col )is not None ][:20 ]
                if not result ['candidates']['cpf']:
                    if any (re .match (r'^\s*\d{3}\.??\d{3}\.??\d{3}-?\d{2}\s*$',s )for s in sample ):
                        result ['candidates']['cpf']=col 
                if not result ['candidates']['email']:
                    if any ('@'in s for s in sample ):
                        result ['candidates']['email']=col 
                if not result ['candidates']['name']:
                    if any (len (s .split ())>=2 for s in sample ):
                        result ['candidates']['name']=col 
            if result ['candidates']['cpf']and result ['candidates']['name']:
                result ['type']='cpf'
            elif result ['candidates']['email']:
                result ['type']='email'
            else :
                result ['type']=None 
            return result 
        except ImportError :
            pass 
    preview =_fallback_read_preview (path ,max_rows =sample_rows )
    result ['preview']=preview 
    cols =preview .get ('columns',[])
    for col in cols :
        sample =[]
        for row in preview .get ('rows',[]):
            v =row .get (col )
            if v is not None :
                sample .append (str (v ))
        for s in sample [:20 ]:
            if re .match (r'^\s*\d{3}\.??\d{3}\.??\d{3}-?\d{2}\s*$',s ):
                result ['candidates']['cpf']=col 
                break 
        if not result ['candidates']['email']:
            if any ('@'in s for s in sample [:20 ]):
                result ['candidates']['email']=col 
        if not result ['candidates']['name']:
            if any (len (s .split ())>=2 for s in sample [:20 ]):
                result ['candidates']['name']=col 
    if result ['candidates']['cpf']and result ['candidates']['name']:
        result ['type']='cpf'
    elif result ['candidates']['email']:
        result ['type']='email'
    else :
        result ['type']=None 
    return result 